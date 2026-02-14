"""
Bulk import students from XLSX file into the database as PENDING registrations.

Usage:
    python scripts/import_students.py "/path/to/sudah siap.xlsx"

XLSX format (header at row 10):
    *Person ID | *Organization | *Person Name | *Gender | Contact | Email | ...

Gender mapping: 1 = Laki-Laki, 2 = Perempuan
Organization = kelas_jurusan (e.g. "X-A", "XI-IPA-1")
"""

import sys
import asyncio
from pathlib import Path

# Add backend root to sys.path so we can import app modules
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from sqlalchemy import select, and_

from app.config.database import engine, async_session_maker
from app.models.user import User
from app.models.siswa_profile import SiswaProfile
from app.enums import UserType, RegistrationStatus, JenisKelamin, StatusSiswa


GENDER_MAP = {
    1: JenisKelamin.laki_laki,
    2: JenisKelamin.perempuan,
}

BATCH_SIZE = 100


async def import_students(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    # Find header row (contains "*Person ID")
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        if row and row[0] and str(row[0]).strip() == "*Person ID":
            header_row_idx = i
            break

    if header_row_idx is None:
        print("ERROR: Could not find header row with '*Person ID'")
        return

    data_start = header_row_idx + 1
    print(f"Header found at row {header_row_idx}, data starts at row {data_start}")

    # Collect rows
    rows = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        person_id = row[0]
        organization = row[1]
        person_name = row[2]
        gender = row[3]

        if not person_name:
            continue

        nama = str(person_name).strip()
        kelas = str(organization).strip() if organization else ""
        gender_enum = GENDER_MAP.get(gender)
        kontak = str(row[4]).strip() if row[4] else ""

        if not gender_enum:
            print(f"  SKIP (invalid gender={gender}): {nama}")
            continue

        rows.append({
            "nama_lengkap": nama,
            "kelas_jurusan": kelas,
            "jenis_kelamin": gender_enum,
            "kontak": kontak,
        })

    wb.close()
    print(f"Found {len(rows)} valid student rows")

    # Import into database
    async with async_session_maker() as session:
        created = 0
        skipped = 0

        for i, row_data in enumerate(rows):
            # Dedup check: same nama_lengkap + kelas_jurusan already exists
            result = await session.execute(
                select(SiswaProfile).where(
                    and_(
                        SiswaProfile.nama_lengkap == row_data["nama_lengkap"],
                        SiswaProfile.kelas_jurusan == row_data["kelas_jurusan"],
                    )
                )
            )
            if result.scalar_one_or_none():
                skipped += 1
                continue

            # Create User (PENDING, no username/password)
            user = User(
                user_type=UserType.siswa,
                registration_status=RegistrationStatus.pending,
                username=None,
                password_hash=None,
                is_active=False,
            )
            session.add(user)
            await session.flush()

            # Create SiswaProfile (partial)
            profile = SiswaProfile(
                user_id=user.user_id,
                nama_lengkap=row_data["nama_lengkap"],
                jenis_kelamin=row_data["jenis_kelamin"],
                kelas_jurusan=row_data["kelas_jurusan"],
                kontak=row_data["kontak"],
                status_siswa=StatusSiswa.aktif,
                kewarganegaraan="Indonesia",
            )
            session.add(profile)
            created += 1

            # Batch commit
            if created % BATCH_SIZE == 0:
                await session.commit()
                print(f"  Committed batch... ({created} created so far)")

        # Final commit
        await session.commit()
        print(f"\nDone! Created: {created}, Skipped (duplicates): {skipped}")

    await engine.dispose()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python scripts/import_students.py <path-to-xlsx>")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    if not Path(xlsx_path).exists():
        print(f"ERROR: File not found: {xlsx_path}")
        sys.exit(1)

    asyncio.run(import_students(xlsx_path))
